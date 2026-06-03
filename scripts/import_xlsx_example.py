"""Мини-пример импорта справочника partner_types из Excel."""

import os

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import MetaData, Table, create_engine
from sqlalchemy.dialects.postgresql import insert


load_dotenv()


def import_partner_types(path: str = "data.xlsx") -> None:
    url = os.getenv("DATABASE_URL")
    if not url:
        raise RuntimeError("Не задан DATABASE_URL")
    rows = [
        {"name": str(row[0]).strip()}
        for row in load_workbook(path, data_only=True)["partner_types"].iter_rows(min_row=2, values_only=True)
        if row[0]
    ]
    if rows:
        engine = create_engine(url, future=True)
        table = Table("partner_types", MetaData(), autoload_with=engine)
        with engine.begin() as connection:
            connection.execute(insert(table).values(rows).on_conflict_do_nothing(index_elements=["name"]))


if __name__ == "__main__":
    import_partner_types()
